"""MigraciÃ³n a la fuente Ãºnica de verdad (SPEC-007 Fase B).

Reinicia el sistema activo a las 51 vacas del archivo Ãºnico:
  - animales: conserva solo las 51, les agrega etapa_actual y foto_principal
  - produccion_lechera / eventos_reproductivos: recargados del archivo Ãºnico
  - hitos_reproductivos: queda vacÃ­o (su fuente, FICHAS, es obsoleta)
  - pesajes: se conservan SOLO los de las 51 vacas (histÃ³rico congelado)
  - notas_vaca: observaciones de Robin desde el catÃ¡logo VIENTRES
  - celos reales observados -> eventos_reproductivos ('Celo Posparto')

Uso:
    python scripts/migrar_fuente_unica.py
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from etl.config import RAW_DIR, db_config  # noqa: E402
from etl.extract import iter_animal_sheets, load_curva_grid  # noqa: E402
from etl.load import (  # noqa: E402
    load_catalogs,
    load_eventos_reproductivos,
    load_produccion,
)
from etl.transform import parse_produccion_curva, parse_repro_events  # noqa: E402

import os  # noqa: E402
import psycopg2  # noqa: E402


def conectar_como_admin():
    """La migración es mantenimiento: requiere TRUNCATE/ownership (admin)."""
    cfg = dict(db_config())
    cfg["user"] = os.environ.get("PG_SUPER_USER", "agrorden_admin")
    cfg["password"] = os.environ["PG_SUPER_PASSWORD"]
    return psycopg2.connect(**cfg)

ARCHIVO_UNICO = ("7.CURVA DE PRODUCCIÓN - ORGANIZADO - RESPALDO - "
                 "respaldo 20260820_010648.xlsx")
HOJA_VIENTRES = "VIENTRES DISPONIBLES DEL SENA "

# Columnas del catÃ¡logo VIENTRES (verificadas por inspecciÃ³n):
COL_NUMERO, COL_ETAPA, COL_FECHA_PARTO = 2, 3, 4
COL_CELO_REAL, COL_OBSERVACIONES = 10, 12


def leer_vientres(ruta: Path) -> dict[str, dict]:
    from openpyxl import load_workbook
    wb = load_workbook(ruta, data_only=True)
    ws = wb[HOJA_VIENTRES]
    datos: dict[str, dict] = {}
    for fila in range(15, ws.max_row + 1):
        numero = ws.cell(row=fila, column=COL_NUMERO).value
        if numero in (None, ""):
            continue
        datos[str(numero).strip()] = {
            "etapa_actual": ws.cell(row=fila, column=COL_ETAPA).value,
            "celo_real": ws.cell(row=fila, column=COL_CELO_REAL).value,
            "observaciones": ws.cell(row=fila, column=COL_OBSERVACIONES).value,
        }
    wb.close()
    return datos


def main() -> int:
    ruta = RAW_DIR / ARCHIVO_UNICO
    if not ruta.exists():
        print(f"No estÃ¡ el archivo Ãºnico: {ruta}")
        return 1

    refs = list(iter_animal_sheets(ruta))
    numeros_archivo = [ref.numero_base for ref in refs]
    print(f"Archivo Ãºnico: {len(refs)} hojas de vaca")

    vientres = leer_vientres(ruta)
    print(f"CatÃ¡logo VIENTRES: {len(vientres)} vacas")

    conn = conectar_como_admin()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id_interno, numero_visible FROM animales "
                "WHERE numero_visible = ANY(%s)",
                (numeros_archivo,),
            )
            filas = cur.fetchall()
        ids = {numero: id_int for id_int, numero in filas}
        faltantes = set(numeros_archivo) - set(ids)
        if faltantes:
            print(f"AVISO: sin fila en animales: {sorted(faltantes)}")

        mantenidos = list(ids.values())
        cur = conn.cursor()

        # 1) Reiniciar lo que se recarga desde la fuente única
        #    (antes de retirar animales, para no violar llaves foráneas)
        for tabla in ("produccion_lechera", "eventos_reproductivos",
                      "hitos_reproductivos", "notas_vaca", "etl_cuarentena"):
            cur.execute(f"TRUNCATE {tabla} RESTART IDENTITY CASCADE")
        print("tablas operativas reiniciadas")

        # 2) Retirar animales fuera del archivo único (y sus dependencias)
        cur.execute("DELETE FROM pesajes WHERE id_animal != ALL(%s::uuid[])", (mantenidos,))
        print(f"pesajes retirados: {cur.rowcount}")
        for tabla in ("hitos_reproductivos", "eventos_sanitarios"):
            cur.execute(f"DELETE FROM {tabla} WHERE id_animal != ALL(%s::uuid[])",
                        (mantenidos,))
            print(f"{tabla} retiradas: {cur.rowcount}")
        cur.execute("UPDATE animales SET id_madre = NULL "
                    "WHERE id_madre != ALL(%s::uuid[])", (mantenidos,))
        cur.execute("DELETE FROM animales WHERE id_interno != ALL(%s::uuid[])",
                    (mantenidos,))
        print(f"animales retirados: {cur.rowcount}")

        # 3) Recargar producciÃ³n y eventos reproductivos del archivo Ãºnico
        tipos_lotes, tipos_san, tipos_repro = load_catalogs(conn)
        tipo_celo = tipos_repro["Celo Posparto"]
        n_prod = n_repro = n_notas = n_celos = 0
        for ref in refs:
            grid_data = load_curva_grid(ruta, ref.sheet)
            evs, _ = parse_repro_events(
                grid_data["meta"], ref.numero_base, ruta.name, ref.sheet)
            n_repro += load_eventos_reproductivos(
                conn, evs, ids, tipos_repro)
            prods, _ = parse_produccion_curva(
                grid_data["grid"], grid_data["meses"], ref.numero_base,
                ruta.name, ref.sheet)
            n_prod += load_produccion(conn, prods, ids)

            info = vientres.get(ref.numero_base, {})
            etapa = info.get("etapa_actual")
            if etapa not in (None, ""):
                cur.execute(
                    "UPDATE animales SET etapa_actual = %s WHERE id_interno = %s",
                    (str(etapa), ids[ref.numero_base]),
                )
                n_notas += cur.rowcount
            celo = info.get("celo_real")
            if celo not in (None, ""):
                cur.execute(
                    "INSERT INTO eventos_reproductivos "
                    "(id_animal, id_tipo_evento, fecha_evento, provisional) "
                    "VALUES (%s, %s, %s, FALSE)",
                    (ids[ref.numero_base], tipo_celo,
                     celo.date() if hasattr(celo, "date") else celo),
                )
                n_celos += 1
            obs = info.get("observaciones")
            if obs not in (None, ""):
                cur.execute(
                    "INSERT INTO notas_vaca (id_animal, observacion) "
                    "VALUES (%s, %s)",
                    (ids[ref.numero_base], str(obs)),
                )

        # 4) Trazabilidad de captura en los pesajes histÃ³ricos conservados
        cur.execute("UPDATE pesajes SET fuente = 'excel' WHERE fuente IS NULL")

        conn.commit()
        print(f"MigraciÃ³n OK: {len(ids)} vacas activas Â· {n_prod} producciones Â· "
              f"{n_repro} eventos reproductivos Â· {n_celos} celos reales Â· "
              f"{n_notas} etapas actualizadas")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())

