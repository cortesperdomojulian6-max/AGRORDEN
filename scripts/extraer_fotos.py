"""Extrae las fotos de cada vaca del archivo único (SPEC-007 RF-01).

Crea data/fotos/<numero_visible>/foto_1.png, foto_2.png, ...

Uso:
    python scripts/extraer_fotos.py
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from etl.config import RAW_DIR  # noqa: E402
from etl.extract import parse_sheet_identity  # noqa: E402

ARCHIVO_UNICO = ("7.CURVA DE PRODUCCIÓN - ORGANIZADO - RESPALDO - "
                 "respaldo 20260820_010648.xlsx")
DESTINO = Path(__file__).resolve().parent.parent / "data" / "fotos"


def main() -> int:
    from openpyxl import load_workbook

    ruta = RAW_DIR / ARCHIVO_UNICO
    if not ruta.exists():
        print(f"No está el archivo único: {ruta}")
        return 1
    DESTINO.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(ruta, data_only=True)
    extraidas = 0
    vacas = 0
    for nombre in wb.sheetnames:
        animal, _ = parse_sheet_identity(nombre)
        if animal is None:
            continue  # portada y catálogos
        ws = wb[nombre]
        imagenes = getattr(ws, "_images", [])
        if not imagenes:
            continue
        carpeta = DESTINO / animal
        carpeta.mkdir(exist_ok=True)
        for j, img in enumerate(imagenes, start=1):
            destino = carpeta / f"foto_{j}.{img.format or 'png'}"
            destino.write_bytes(img._data())
            extraidas += 1
        vacas += 1
    wb.close()
    print(f"OK: {extraidas} fotos de {vacas} vacas en {DESTINO}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
